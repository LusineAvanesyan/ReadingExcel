import openpyxl

# No ProductName	ProductID	Price	Comapany
# 1	varung	112	300	A company
# 2	lolik	115	350	B company
# 3	ddum	124	110	B company
# 4	grich	255	140	B company
# 5	tetr	800	310	A company

file = openpyxl.load_workbook("inventory.xlsx")

products = file["Sheet1"]
company_name_count = {}

for product_row in range(2, products.max_row + 1):
    companyName = products.cell(product_row, 5).value
    print(companyName)

    if companyName in company_name_count:
        count = company_name_count[companyName]
        company_name_count[companyName] = count + 1 
    else:
        company_name_count[companyName] = 1


print("hello", company_name_count)
